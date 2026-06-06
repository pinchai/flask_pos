import os
import time
from flask import render_template, redirect, url_for, request, flash, abort, current_app
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from extensions import db
from models import User, Shop, Category, Product, PaymentMethod, Sale, SaleItem
from . import dashboard_bp

@dashboard_bp.route('/')
def hello_world():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard.admin_index'))
    return redirect(url_for('dashboard.login'))

@dashboard_bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard.admin_index'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = True if request.form.get('remember') else False

        user = User.query.filter_by(username=username).first()

        if not user or not user.check_password(password):
            flash('Please check your login details and try again.', 'danger')
            return redirect(url_for('dashboard.login'))

        if user.status != 'approve':
            flash('Your account is not approved yet.', 'warning')
            return redirect(url_for('dashboard.login'))

        login_user(user, remember=remember)
        return redirect(url_for('dashboard.admin_index'))

    return render_template('admin/login.html')

@dashboard_bp.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have logged out.', 'success')
    return redirect(url_for('dashboard.login'))

@dashboard_bp.route('/admin')
@login_required
def admin_index():
    from datetime import datetime, timedelta
    from sqlalchemy import func

    user_count = User.query.count()
    shop_count = Shop.query.filter_by(user_id=current_user.id).count()
    product_count = Product.query.filter_by(user_id=current_user.id).count()
    sale_count = Sale.query.filter_by(user_id=current_user.id).count()

    # 1. 15-day revenue trend query
    fifteen_days_ago = datetime.utcnow() - timedelta(days=15)
    daily_sales = db.session.query(
        func.date(Sale.sale_date).label('day'),
        func.sum(Sale.total).label('revenue')
    ).filter(Sale.sale_date >= fifteen_days_ago, Sale.user_id == current_user.id)\
     .group_by(func.date(Sale.sale_date))\
     .order_by('day').all()

    # Create mapping of day string -> revenue float
    sales_map = {str(day): float(revenue) for day, revenue in daily_sales}
    revenue_chart_data = []
    for i in range(14, -1, -1):
        day_dt = datetime.now() - timedelta(days=i)
        day_str = day_dt.strftime('%Y-%m-%d')
        revenue_chart_data.append({
            'date': day_dt.strftime('%b %d'),
            'revenue': sales_map.get(day_str, 0.0)
        })

    # 2. Product distribution by Category query (scoped to logged-in user)
    category_share = db.session.query(
        Category.name,
        func.count(Product.id)
    ).outerjoin(Product, (Product.category_id == Category.id) & (Product.user_id == current_user.id))\
     .filter(Category.user_id == current_user.id)\
     .group_by(Category.id).all()

    category_chart_data = [
        {'category': name, 'count': count}
        for name, count in category_share
    ]

    # 3. Revenue by Shop Location query
    shop_sales_data = db.session.query(
        Shop.name,
        func.sum(Sale.total)
    ).join(Sale, Sale.shop_id == Shop.id)\
     .filter(Sale.user_id == current_user.id)\
     .group_by(Shop.id).all()

    shop_chart_data = [
        {'shop': name, 'total': float(total) if total else 0.0}
        for name, total in shop_sales_data
    ]

    return render_template(
        'admin/dashboard.html',
        user_count=user_count,
        shop_count=shop_count,
        product_count=product_count,
        sale_count=sale_count,
        revenue_chart=revenue_chart_data,
        category_chart=category_chart_data,
        shop_chart=shop_chart_data
    )

# --- USER CRUD ---
@dashboard_bp.route('/admin/users')
@login_required
def list_users():
    if current_user.type != 'admin':
        flash('Access denied. Admin role required.', 'danger')
        return redirect(url_for('dashboard.admin_index'))
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    query = User.query
    if search:
        query = query.filter(
            User.username.ilike(f'%{search}%') |
            User.type.ilike(f'%{search}%') |
            User.status.ilike(f'%{search}%')
        )
    pagination = query.paginate(page=page, per_page=10, error_out=False)
    return render_template('admin/users/list.html', users=pagination.items, pagination=pagination, search=search)

@dashboard_bp.route('/admin/users/create', methods=['GET', 'POST'])
@login_required
def create_user():
    if current_user.type != 'admin':
        flash('Access denied. Admin role required.', 'danger')
        return redirect(url_for('dashboard.admin_index'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        status = request.form.get('status', 'pending')
        utype = request.form.get('type', 'student')

        profile_path = None
        profile_file = request.files.get('profile')
        if profile_file and profile_file.filename != '':
            filename = secure_filename(profile_file.filename)
            filename = f"{int(time.time())}_{filename}"
            upload_dir = os.path.join(current_app.static_folder, 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            profile_file.save(os.path.join(upload_dir, filename))
            profile_path = f"/static/uploads/{filename}"

        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
            return redirect(url_for('dashboard.create_user'))

        new_user = User(username=username, profile=profile_path, status=status, type=utype)
        new_user.set_password(password)

        db.session.add(new_user)
        db.session.commit()

        flash('User created successfully!', 'success')
        return redirect(url_for('dashboard.list_users'))

    return render_template('admin/users/form.html', user=None)

@dashboard_bp.route('/admin/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if current_user.type != 'admin':
        flash('Access denied. Admin role required.', 'danger')
        return redirect(url_for('dashboard.admin_index'))
    user = User.query.get_or_404(user_id)

    if request.method == 'POST':
        user.username = request.form.get('username')
        user.status = request.form.get('status')
        user.type = request.form.get('type')

        profile_file = request.files.get('profile')
        if profile_file and profile_file.filename != '':
            filename = secure_filename(profile_file.filename)
            filename = f"{int(time.time())}_{filename}"
            upload_dir = os.path.join(current_app.static_folder, 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            profile_file.save(os.path.join(upload_dir, filename))
            user.profile = f"/static/uploads/{filename}"

        password = request.form.get('password')
        if password:
            user.set_password(password)

        db.session.commit()
        flash('User updated successfully!', 'success')
        return redirect(url_for('dashboard.list_users'))

    return render_template('admin/users/form.html', user=user)

@dashboard_bp.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if current_user.type != 'admin':
        flash('Access denied. Admin role required.', 'danger')
        return redirect(url_for('dashboard.admin_index'))
    if user_id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('dashboard.list_users'))

    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash('User deleted successfully!', 'success')
    return redirect(url_for('dashboard.list_users'))


# --- SHOP CRUD ---
@dashboard_bp.route('/admin/shops')
@login_required
def list_shops():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    query = Shop.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(
            Shop.name.ilike(f'%{search}%') |
            Shop.address.ilike(f'%{search}%') |
            Shop.description.ilike(f'%{search}%')
        )
    pagination = query.paginate(page=page, per_page=10, error_out=False)
    return render_template('admin/shops/list.html', shops=pagination.items, pagination=pagination, search=search)

@dashboard_bp.route('/admin/shops/create', methods=['GET', 'POST'])
@login_required
def create_shop():
    users = User.query.all()
    if request.method == 'POST':
        name = request.form.get('name')
        address = request.form.get('address')
        description = request.form.get('description')
        user_id = current_user.id

        logo_path = None
        logo_file = request.files.get('logo')
        if logo_file and logo_file.filename != '':
            filename = secure_filename(logo_file.filename)
            filename = f"{int(time.time())}_{filename}"
            upload_dir = os.path.join(current_app.static_folder, 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            logo_file.save(os.path.join(upload_dir, filename))
            logo_path = f"/static/uploads/{filename}"

        new_shop = Shop(name=name, address=address, logo=logo_path, description=description, user_id=user_id)
        db.session.add(new_shop)
        db.session.commit()

        flash('Shop created successfully!', 'success')
        return redirect(url_for('dashboard.list_shops'))

    return render_template('admin/shops/form.html', shop=None, users=users)

@dashboard_bp.route('/admin/shops/edit/<int:shop_id>', methods=['GET', 'POST'])
@login_required
def edit_shop(shop_id):
    shop = Shop.query.filter_by(id=shop_id, user_id=current_user.id).first_or_404()
    users = User.query.all()

    if request.method == 'POST':
        shop.name = request.form.get('name')
        shop.address = request.form.get('address')
        shop.description = request.form.get('description')
        shop.user_id = current_user.id

        logo_file = request.files.get('logo')
        if logo_file and logo_file.filename != '':
            filename = secure_filename(logo_file.filename)
            filename = f"{int(time.time())}_{filename}"
            upload_dir = os.path.join(current_app.static_folder, 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            logo_file.save(os.path.join(upload_dir, filename))
            shop.logo = f"/static/uploads/{filename}"

        db.session.commit()
        flash('Shop updated successfully!', 'success')
        return redirect(url_for('dashboard.list_shops'))

    return render_template('admin/shops/form.html', shop=shop, users=users)

@dashboard_bp.route('/admin/shops/delete/<int:shop_id>', methods=['POST'])
@login_required
def delete_shop(shop_id):
    shop = Shop.query.filter_by(id=shop_id, user_id=current_user.id).first_or_404()
    db.session.delete(shop)
    db.session.commit()
    flash('Shop deleted successfully!', 'success')
    return redirect(url_for('dashboard.list_shops'))


# --- CATEGORY CRUD ---
@dashboard_bp.route('/admin/categories')
@login_required
def list_categories():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    query = Category.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(
            Category.name.ilike(f'%{search}%') |
            Category.remark.ilike(f'%{search}%')
        )
    pagination = query.paginate(page=page, per_page=10, error_out=False)
    return render_template('admin/categories/list.html', categories=pagination.items, pagination=pagination, search=search)

@dashboard_bp.route('/admin/categories/create', methods=['GET', 'POST'])
@login_required
def create_category():
    users = User.query.all()
    if request.method == 'POST':
        name = request.form.get('name')
        remark = request.form.get('remark')
        user_id = current_user.id

        new_cat = Category(name=name, remark=remark, user_id=user_id)
        db.session.add(new_cat)
        db.session.commit()

        flash('Category created successfully!', 'success')
        return redirect(url_for('dashboard.list_categories'))

    return render_template('admin/categories/form.html', category=None, users=users)

@dashboard_bp.route('/admin/categories/edit/<int:category_id>', methods=['GET', 'POST'])
@login_required
def edit_category(category_id):
    category = Category.query.filter_by(id=category_id, user_id=current_user.id).first_or_404()
    users = User.query.all()

    if request.method == 'POST':
        category.name = request.form.get('name')
        category.remark = request.form.get('remark')
        category.user_id = current_user.id

        db.session.commit()
        flash('Category updated successfully!', 'success')
        return redirect(url_for('dashboard.list_categories'))

    return render_template('admin/categories/form.html', category=category, users=users)

@dashboard_bp.route('/admin/categories/delete/<int:category_id>', methods=['POST'])
@login_required
def delete_category(category_id):
    category = Category.query.filter_by(id=category_id, user_id=current_user.id).first_or_404()
    db.session.delete(category)
    db.session.commit()
    flash('Category deleted successfully!', 'success')
    return redirect(url_for('dashboard.list_categories'))


# --- PRODUCT CRUD ---
@dashboard_bp.route('/admin/products')
@login_required
def list_products():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    query = Product.query.filter_by(user_id=current_user.id)
    if search:
        query = query.join(Category).filter(
            Product.name.ilike(f'%{search}%') |
            Product.remark.ilike(f'%{search}%') |
            Category.name.ilike(f'%{search}%')
        )
    pagination = query.paginate(page=page, per_page=10, error_out=False)
    return render_template('admin/products/list.html', products=pagination.items, pagination=pagination, search=search)

@dashboard_bp.route('/admin/products/create', methods=['GET', 'POST'])
@login_required
def create_product():
    categories = Category.query.filter_by(user_id=current_user.id).all()
    users = User.query.all()

    if not categories:
        flash('Please create a category first before adding products.', 'warning')
        return redirect(url_for('dashboard.list_categories'))

    if request.method == 'POST':
        name = request.form.get('name')
        category_id = request.form.get('category_id')
        cost = float(request.form.get('cost') or 0.0)
        price = float(request.form.get('price') or 0.0)
        stock = float(request.form.get('stock') or 0.0)
        remark = request.form.get('remark')
        user_id = current_user.id

        image_path = None
        image_file = request.files.get('image')
        if image_file and image_file.filename != '':
            filename = secure_filename(image_file.filename)
            filename = f"{int(time.time())}_{filename}"
            upload_dir = os.path.join(current_app.static_folder, 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            image_file.save(os.path.join(upload_dir, filename))
            image_path = f"/static/uploads/{filename}"

        new_prod = Product(
            name=name, category_id=category_id, cost=cost, price=price,
            stock=stock, image=image_path, remark=remark, user_id=user_id
        )
        db.session.add(new_prod)
        db.session.commit()

        flash('Product created successfully!', 'success')
        return redirect(url_for('dashboard.list_products'))

    return render_template('admin/products/form.html', product=None, categories=categories, users=users)

@dashboard_bp.route('/admin/products/edit/<int:product_id>', methods=['GET', 'POST'])
@login_required
def edit_product(product_id):
    product = Product.query.filter_by(id=product_id, user_id=current_user.id).first_or_404()
    categories = Category.query.filter_by(user_id=current_user.id).all()
    users = User.query.all()

    if request.method == 'POST':
        product.name = request.form.get('name')
        product.category_id = request.form.get('category_id')
        product.cost = float(request.form.get('cost') or 0.0)
        product.price = float(request.form.get('price') or 0.0)
        product.stock = float(request.form.get('stock') or 0.0)
        product.remark = request.form.get('remark')
        product.user_id = current_user.id

        image_file = request.files.get('image')
        if image_file and image_file.filename != '':
            filename = secure_filename(image_file.filename)
            filename = f"{int(time.time())}_{filename}"
            upload_dir = os.path.join(current_app.static_folder, 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            image_file.save(os.path.join(upload_dir, filename))
            product.image = f"/static/uploads/{filename}"

        db.session.commit()
        flash('Product updated successfully!', 'success')
        return redirect(url_for('dashboard.list_products'))

    return render_template('admin/products/form.html', product=product, categories=categories, users=users)

@dashboard_bp.route('/admin/products/delete/<int:product_id>', methods=['POST'])
@login_required
def delete_product(product_id):
    product = Product.query.filter_by(id=product_id, user_id=current_user.id).first_or_404()
    db.session.delete(product)
    db.session.commit()
    flash('Product deleted successfully!', 'success')
    return redirect(url_for('dashboard.list_products'))


# --- PAYMENT METHOD CRUD ---
@dashboard_bp.route('/admin/payment-methods')
@login_required
def list_payment_methods():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    query = PaymentMethod.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(
            PaymentMethod.name.ilike(f'%{search}%') |
            PaymentMethod.remark.ilike(f'%{search}%')
        )
    pagination = query.paginate(page=page, per_page=10, error_out=False)
    return render_template('admin/payment_methods/list.html', payment_methods=pagination.items, pagination=pagination, search=search)

@dashboard_bp.route('/admin/payment-methods/create', methods=['GET', 'POST'])
@login_required
def create_payment_method():
    users = User.query.all()
    if request.method == 'POST':
        name = request.form.get('name')
        remark = request.form.get('remark')
        user_id = current_user.id

        new_pm = PaymentMethod(name=name, remark=remark, user_id=user_id)
        db.session.add(new_pm)
        db.session.commit()

        flash('Payment Method created successfully!', 'success')
        return redirect(url_for('dashboard.list_payment_methods'))

    return render_template('admin/payment_methods/form.html', payment_method=None, users=users)

@dashboard_bp.route('/admin/payment-methods/edit/<int:pm_id>', methods=['GET', 'POST'])
@login_required
def edit_payment_method(pm_id):
    payment_method = PaymentMethod.query.filter_by(id=pm_id, user_id=current_user.id).first_or_404()
    users = User.query.all()

    if request.method == 'POST':
        payment_method.name = request.form.get('name')
        payment_method.remark = request.form.get('remark')
        payment_method.user_id = current_user.id

        db.session.commit()
        flash('Payment Method updated successfully!', 'success')
        return redirect(url_for('dashboard.list_payment_methods'))

    return render_template('admin/payment_methods/form.html', payment_method=payment_method, users=users)

@dashboard_bp.route('/admin/payment-methods/delete/<int:pm_id>', methods=['POST'])
@login_required
def delete_payment_method(pm_id):
    payment_method = PaymentMethod.query.filter_by(id=pm_id, user_id=current_user.id).first_or_404()
    db.session.delete(payment_method)
    db.session.commit()
    flash('Payment Method deleted successfully!', 'success')
    return redirect(url_for('dashboard.list_payment_methods'))


# --- SALE CRUD ---
@dashboard_bp.route('/admin/sales')
@login_required
def list_sales():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    query = Sale.query.filter_by(user_id=current_user.id)
    if search:
        filter_cond = Shop.name.ilike(f'%{search}%') | PaymentMethod.name.ilike(f'%{search}%') | User.username.ilike(f'%{search}%')
        if search.isdigit():
            filter_cond = filter_cond | (Sale.id == int(search))
        query = query.join(Shop).join(PaymentMethod).join(User).filter(filter_cond)
    pagination = query.order_by(Sale.sale_date.desc()).paginate(page=page, per_page=10, error_out=False)
    return render_template('admin/sales/list.html', sales=pagination.items, pagination=pagination, search=search)

@dashboard_bp.route('/admin/sales/view/<int:sale_id>')
@login_required
def view_sale(sale_id):
    sale = Sale.query.filter_by(id=sale_id, user_id=current_user.id).first_or_404()
    return render_template('admin/sales/view.html', sale=sale)

@dashboard_bp.route('/admin/sales/view/<int:sale_id>/pdf')
@login_required
def view_sale_pdf(sale_id):
    import io
    from flask import send_file
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    sale = Sale.query.filter_by(id=sale_id, user_id=current_user.id).first_or_404()
    
    buffer = io.BytesIO()
    
    # Page setup
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom paragraph styles
    title_style = ParagraphStyle(
        'InvoiceTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=22,
        leading=26,
        textColor=colors.HexColor('#1F4E78')
    )
    
    section_title_style = ParagraphStyle(
        'InvoiceSectionTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=15,
        textColor=colors.HexColor('#1F4E78')
    )
    
    normal_bold = ParagraphStyle(
        'InvoiceNormalBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=13
    )
    
    normal_text = ParagraphStyle(
        'InvoiceNormalText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=13
    )
    
    right_text = ParagraphStyle(
        'InvoiceRightText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=13,
        alignment=2
    )
    
    right_bold = ParagraphStyle(
        'InvoiceRightBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=13,
        alignment=2
    )
    
    header_bold_left = ParagraphStyle(
        'InvoiceHeaderBoldLeft',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=13,
        textColor=colors.white
    )
    
    header_bold_right = ParagraphStyle(
        'InvoiceHeaderBoldRight',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=13,
        textColor=colors.white,
        alignment=2
    )
    
    # 1. Header (Title and Date)
    header_data = [
        [
            Paragraph("SALE INVOICE", title_style),
            Paragraph(f"Date-Time: {sale.sale_date.strftime('%Y-%m-%d %H:%M:%S')}", right_text)
        ]
    ]
    header_table = Table(header_data, colWidths=[250, 250])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 10))
    
    # 2. Meta Info Card
    meta_cols = [
        [
            Paragraph("<b>Shop Info</b>", section_title_style),
            Paragraph("<b>Salesperson</b>", section_title_style),
            Paragraph("<b>Payment details</b>", section_title_style)
        ],
        [
            Paragraph(f"<b>{sale.shop.name}</b><br/>Address: {sale.shop.address}", normal_text),
            Paragraph(f"<b>{sale.user.username}</b><br/>Type: {sale.user.type or '-'}", normal_text),
            Paragraph(f"Method: {sale.payment_method.name}<br/>Discount: {sale.discount_pct}%", normal_text)
        ]
    ]
    meta_table = Table(meta_cols, colWidths=[166, 166, 168])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,0), 4),
        ('BOTTOMPADDING', (0,1), (-1,-1), 10),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F2F2F2')),
        ('TOPPADDING', (0,0), (-1,0), 4),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#BFBFBF')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 15))
    
    # 3. Item List Title
    story.append(Paragraph("<b>Transaction Items</b>", section_title_style))
    story.append(Spacer(1, 6))
    
    # 4. Item List Table
    table_data = [
        [
            Paragraph("Product", header_bold_left),
            Paragraph("Price", header_bold_right),
            Paragraph("Qty", header_bold_right),
            Paragraph("Subtotal", header_bold_right)
        ]
    ]
    
    for item in sale.items:
        subtotal = float(item.price or 0.0) * int(item.qty or 0)
        table_data.append([
            Paragraph(item.product.name, normal_text),
            Paragraph(f"${float(item.price or 0.0):.2f}", right_text),
            Paragraph(str(item.qty), right_text),
            Paragraph(f"${subtotal:.2f}", right_text)
        ])
        
    items_table = Table(table_data, colWidths=[250, 80, 70, 100])
    items_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#1F4E78')),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 15))
    
    # 5. Summary block
    gross = float(sale.total or 0.0) + float(sale.discount_amount or 0.0)
    summary_data = [
        [Paragraph("Gross Subtotal:", normal_text), Paragraph(f"${gross:.2f}", right_text)],
        [Paragraph(f"Discount ({sale.discount_pct}%):", normal_text), Paragraph(f"-${float(sale.discount_amount or 0.0):.2f}", right_text)],
        [Paragraph("<b>Total Due:</b>", normal_bold), Paragraph(f"<b>${float(sale.total or 0.0):.2f}</b>", right_bold)],
        [Paragraph("Amount Paid:", normal_text), Paragraph(f"${float(sale.paid_amount or 0.0):.2f}", right_text)]
    ]
    summary_table = Table(summary_data, colWidths=[150, 100])
    summary_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('LINEBELOW', (0,1), (1,1), 0.5, colors.HexColor('#BFBFBF')),
        ('LINEBELOW', (0,2), (1,2), 1.2, colors.HexColor('#1F4E78')),
    ]))
    
    outer_table_data = [
        ["", summary_table]
    ]
    outer_table = Table(outer_table_data, colWidths=[250, 250])
    outer_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (1,0), (1,0), 'RIGHT'),
    ]))
    story.append(outer_table)
    
    # Build PDF
    doc.build(story)
    
    buffer.seek(0)
    filename = f"invoice_sale_{sale.id}.pdf"
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=filename
    )

@dashboard_bp.route('/admin/sales/create', methods=['GET', 'POST'])
@login_required
def create_sale():
    shops = Shop.query.filter_by(user_id=current_user.id).all()
    users = User.query.all()
    products = Product.query.filter_by(user_id=current_user.id).all()
    payment_methods = PaymentMethod.query.filter_by(user_id=current_user.id).all()

    if not shops or not products or not payment_methods:
        flash('Please ensure you have created at least one Shop, Product, and Payment Method before recording a sale.', 'warning')
        return redirect(url_for('dashboard.admin_index'))

    if request.method == 'POST':
        shop_id = request.form.get('shop_id')
        user_id = current_user.id
        payment_method_id = request.form.get('payment_method_id')
        discount_pct = int(request.form.get('discount_pct') or 0)
        discount_amount = float(request.form.get('discount_amount') or 0.0)
        total = float(request.form.get('total') or 0.0)
        paid_amount = float(request.form.get('paid_amount') or 0.0)

        # Create Sale object
        new_sale = Sale(
            shop_id=shop_id, user_id=user_id, payment_method_id=payment_method_id,
            discount_pct=discount_pct, discount_amount=discount_amount,
            total=total, paid_amount=paid_amount
        )
        db.session.add(new_sale)
        
        # Read list of products and quantities submitted dynamically
        product_ids = request.form.getlist('product_ids[]')
        qtys = request.form.getlist('qtys[]')

        for prod_id, qty_val in zip(product_ids, qtys):
            qty = int(qty_val)
            product = Product.query.get(prod_id)
            if product:
                # Deduct stock level
                product.stock = float(product.stock) - float(qty)
                
                # Create SaleItem
                item = SaleItem(
                    user_id=user_id,
                    sale=new_sale,
                    product_id=prod_id,
                    qty=qty,
                    price=product.price
                )
                db.session.add(item)

        db.session.commit()
        flash('Sale created successfully!', 'success')
        return redirect(url_for('dashboard.list_sales'))

    return render_template(
        'admin/sales/create.html',
        shops=shops, users=users, products=products, payment_methods=payment_methods
    )

@dashboard_bp.route('/admin/sales/delete/<int:sale_id>', methods=['POST'])
@login_required
def delete_sale(sale_id):
    sale = Sale.query.filter_by(id=sale_id, user_id=current_user.id).first_or_404()
    
    # Delete sale items first
    for item in sale.items:
        # Revert stock deduction when deleting a sale
        item.product.stock = float(item.product.stock) + float(item.qty)
        db.session.delete(item)
        
    db.session.delete(sale)
    db.session.commit()
    
    flash('Sale and associated items deleted successfully, inventory levels restored.', 'success')
    return redirect(url_for('dashboard.list_sales'))


# --- SALES REPORT ---
@dashboard_bp.route('/admin/reports/sales')
@login_required
def sales_report():
    from datetime import datetime, timedelta
    from sqlalchemy import func
    
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    branch_id = request.args.get('branch_id', type=int)
    category_id = request.args.get('category_id', type=int)
    product_id = request.args.get('product_id', type=int)

    # Default to last 30 days if not specified
    if not from_date_str:
        from_date_dt = datetime.now() - timedelta(days=30)
        from_date_str = from_date_dt.strftime('%Y-%m-%dT%H:%M')
    else:
        try:
            from_date_dt = datetime.fromisoformat(from_date_str)
        except ValueError:
            from_date_dt = datetime.now() - timedelta(days=30)
            from_date_str = from_date_dt.strftime('%Y-%m-%dT%H:%M')

    if not to_date_str:
        to_date_dt = datetime.now()
        to_date_str = to_date_dt.strftime('%Y-%m-%dT%H:%M')
    else:
        try:
            to_date_dt = datetime.fromisoformat(to_date_str)
        except ValueError:
            to_date_dt = datetime.now()
            to_date_str = to_date_dt.strftime('%Y-%m-%dT%H:%M')

    # Query matching sales
    sales_query = Sale.query.filter(
        Sale.sale_date >= from_date_dt,
        Sale.sale_date <= to_date_dt,
        Sale.user_id == current_user.id
    )

    # Apply additional filters
    if branch_id:
        sales_query = sales_query.filter(Sale.shop_id == branch_id)
    if category_id:
        sale_ids_query = db.session.query(SaleItem.sale_id).join(Product).filter(
            Product.category_id == category_id,
            SaleItem.user_id == current_user.id
        )
        sales_query = sales_query.filter(Sale.id.in_(sale_ids_query))
    if product_id:
        sale_ids_query = db.session.query(SaleItem.sale_id).filter(
            SaleItem.product_id == product_id,
            SaleItem.user_id == current_user.id
        )
        sales_query = sales_query.filter(Sale.id.in_(sale_ids_query))

    # Calculate Aggregates
    sales_subquery = sales_query.with_entities(Sale.id)

    total_sales_val = db.session.query(func.sum(Sale.total)).filter(Sale.id.in_(sales_subquery)).scalar() or 0.0
    total_discount_val = db.session.query(func.sum(Sale.discount_amount)).filter(Sale.id.in_(sales_subquery)).scalar() or 0.0
    total_transactions = sales_query.count()

    # Sales by Shop
    shop_sales = db.session.query(
        Shop.name,
        func.sum(Sale.total).label('shop_total')
    ).join(Sale, Sale.shop_id == Shop.id)\
     .filter(Sale.id.in_(sales_subquery), Shop.user_id == current_user.id)\
     .group_by(Shop.id).all()

    # Sales by Payment Method
    pm_sales = db.session.query(
        PaymentMethod.name,
        func.sum(Sale.total).label('pm_total')
    ).join(Sale, Sale.payment_method_id == PaymentMethod.id)\
     .filter(Sale.id.in_(sales_subquery), PaymentMethod.user_id == current_user.id)\
     .group_by(PaymentMethod.id).all()

    # Get filter dropdown list items scoped to the user
    shops = Shop.query.filter_by(user_id=current_user.id).order_by(Shop.name).all()
    categories = Category.query.filter_by(user_id=current_user.id).order_by(Category.name).all()
    products = Product.query.filter_by(user_id=current_user.id).order_by(Product.name).all()

    # Paginated detailed list of sales
    page = request.args.get('page', 1, type=int)
    pagination = sales_query.order_by(Sale.sale_date.desc()).paginate(page=page, per_page=10, error_out=False)

    return render_template(
        'admin/reports/sales.html',
        sales=pagination.items,
        pagination=pagination,
        total_sales=total_sales_val,
        total_discount=total_discount_val,
        total_transactions=total_transactions,
        shop_sales=shop_sales,
        pm_sales=pm_sales,
        from_date=from_date_str,
        to_date=to_date_str,
        shops=shops,
        categories=categories,
        products=products,
        branch_id=branch_id,
        category_id=category_id,
        product_id=product_id
    )


# --- POS ROUTING & API ---
@dashboard_bp.route('/admin/pos')
@login_required
def pos_index():
    shops = Shop.query.filter_by(user_id=current_user.id).all()
    payment_methods = PaymentMethod.query.filter_by(user_id=current_user.id).all()
    return render_template('admin/pos.html', shops=shops, payment_methods=payment_methods)


@dashboard_bp.route('/admin/api/products')
@login_required
def api_list_products():
    products = Product.query.filter_by(user_id=current_user.id).all()
    out = []
    for p in products:
        out.append({
            'id': p.id,
            'title': p.name,
            'price': float(p.price) if p.price else 0.0,
            'category': p.category.name,
            'image': p.image or f"https://placehold.co/170x170?text={p.name}",
            'stock': float(p.stock) if p.stock else 0.0
        })
    return {'products': out}


@dashboard_bp.route('/admin/api/sales/create', methods=['POST'])
@login_required
def api_create_sale():
    data = request.get_json()
    if not data:
        return {'error': 'No data provided'}, 400
    
    shop_id = data.get('shop_id')
    payment_method_id = data.get('payment_method_id')
    discount_pct = int(data.get('discount_pct') or 0)
    discount_amount = float(data.get('discount_amount') or 0.0)
    total = float(data.get('total') or 0.0)
    paid_amount = float(data.get('paid_amount') or 0.0)
    items = data.get('items', [])
    
    if not shop_id or not payment_method_id or not items:
        return {'error': 'Missing required fields or items'}, 400
        
    new_sale = Sale(
        shop_id=shop_id,
        user_id=current_user.id,
        payment_method_id=payment_method_id,
        discount_pct=discount_pct,
        discount_amount=discount_amount,
        total=total,
        paid_amount=paid_amount
    )
    db.session.add(new_sale)
    
    for item_data in items:
        prod_id = item_data.get('product_id')
        qty = int(item_data.get('qty') or 1)
        product = Product.query.get(prod_id)
        if product:
            product.stock = float(product.stock) - float(qty)
            item = SaleItem(
                user_id=current_user.id,
                sale=new_sale,
                product_id=prod_id,
                qty=qty,
                price=product.price
            )
            db.session.add(item)
            
    db.session.commit()
    return {'success': True, 'sale_id': new_sale.id}


# --- INCOME REPORT ---
@dashboard_bp.route('/admin/reports/income')
@login_required
def income_report():
    from datetime import datetime, timedelta
    from sqlalchemy import func
    
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')

    # Default to last 30 days if not specified
    if not from_date_str:
        from_date_dt = datetime.now() - timedelta(days=30)
        from_date_str = from_date_dt.strftime('%Y-%m-%dT%H:%M')
    else:
        try:
            from_date_dt = datetime.fromisoformat(from_date_str)
        except ValueError:
            from_date_dt = datetime.now() - timedelta(days=30)
            from_date_str = from_date_dt.strftime('%Y-%m-%dT%H:%M')

    if not to_date_str:
        to_date_dt = datetime.now()
        to_date_str = to_date_dt.strftime('%Y-%m-%dT%H:%M')
    else:
        try:
            to_date_dt = datetime.fromisoformat(to_date_str)
        except ValueError:
            to_date_dt = datetime.now()
            to_date_str = to_date_dt.strftime('%Y-%m-%dT%H:%M')

    # Query matching sales
    sales_query = Sale.query.filter(
        Sale.sale_date >= from_date_dt,
        Sale.sale_date <= to_date_dt,
        Sale.user_id == current_user.id
    )
    sales_subquery = sales_query.with_entities(Sale.id)

    # Calculate Aggregates
    total_revenue = db.session.query(func.sum(Sale.total)).filter(Sale.id.in_(sales_subquery)).scalar() or 0.0
    total_revenue = float(total_revenue)
    
    total_cogs = db.session.query(
        func.sum(SaleItem.qty * func.coalesce(Product.cost, 0.0))
    ).join(Product, SaleItem.product_id == Product.id)\
     .filter(SaleItem.sale_id.in_(sales_subquery), Product.user_id == current_user.id, SaleItem.user_id == current_user.id).scalar() or 0.0
    total_cogs = float(total_cogs)

    net_profit = total_revenue - total_cogs
    profit_margin = (net_profit / total_revenue) * 100 if total_revenue > 0 else 0.0

    # 1. Sales by Shop Location
    shop_revenues = db.session.query(
        Shop.id,
        Shop.name,
        func.sum(Sale.total).label('revenue')
    ).join(Sale, Sale.shop_id == Shop.id)\
     .filter(Sale.id.in_(sales_subquery), Shop.user_id == current_user.id)\
     .group_by(Shop.id).all()

    shop_cogs = db.session.query(
        Sale.shop_id,
        func.sum(SaleItem.qty * func.coalesce(Product.cost, 0.0)).label('cogs')
    ).join(SaleItem, SaleItem.sale_id == Sale.id)\
     .join(Product, SaleItem.product_id == Product.id)\
     .filter(Sale.id.in_(sales_subquery), Product.user_id == current_user.id, SaleItem.user_id == current_user.id)\
     .group_by(Sale.shop_id).all()

    cogs_map = {row.shop_id: float(row.cogs or 0.0) for row in shop_cogs}
    shop_breakdown = []
    for row in shop_revenues:
        rev = float(row.revenue or 0.0)
        cogs = cogs_map.get(row.id, 0.0)
        prof = rev - cogs
        margin = (prof / rev) * 100 if rev > 0 else 0.0
        shop_breakdown.append({
            'name': row.name,
            'revenue': rev,
            'cogs': cogs,
            'profit': prof,
            'margin': margin
        })

    # 2. Sales by Category
    category_data = db.session.query(
        Category.name,
        func.sum(SaleItem.qty * SaleItem.price * (1 - func.coalesce(Sale.discount_pct, 0) / 100.0)).label('revenue'),
        func.sum(SaleItem.qty * func.coalesce(Product.cost, 0.0)).label('cogs')
    ).select_from(SaleItem)\
     .join(Sale, SaleItem.sale_id == Sale.id)\
     .join(Product, SaleItem.product_id == Product.id)\
     .join(Category, Product.category_id == Category.id)\
     .filter(Sale.id.in_(sales_subquery), Product.user_id == current_user.id, Category.user_id == current_user.id, SaleItem.user_id == current_user.id)\
     .group_by(Category.id).all()

    category_breakdown = []
    for row in category_data:
        rev = float(row.revenue or 0.0)
        cogs = float(row.cogs or 0.0)
        prof = rev - cogs
        margin = (prof / rev) * 100 if rev > 0 else 0.0
        category_breakdown.append({
            'name': row.name,
            'revenue': rev,
            'cogs': cogs,
            'profit': prof,
            'margin': margin
        })

    # 3. Paginated detailed list of sales
    page = request.args.get('page', 1, type=int)
    pagination = sales_query.order_by(Sale.sale_date.desc()).paginate(page=page, per_page=10, error_out=False)

    sale_ids = [s.id for s in pagination.items]
    cogs_per_sale = {}
    if sale_ids:
        cogs_rows = db.session.query(
            SaleItem.sale_id,
            func.sum(SaleItem.qty * func.coalesce(Product.cost, 0.0)).label('cogs')
        ).join(Product, SaleItem.product_id == Product.id)\
         .filter(SaleItem.sale_id.in_(sale_ids), Product.user_id == current_user.id, SaleItem.user_id == current_user.id)\
         .group_by(SaleItem.sale_id).all()
        cogs_per_sale = {row.sale_id: float(row.cogs or 0.0) for row in cogs_rows}

    detailed_sales = []
    for sale in pagination.items:
        rev = float(sale.total)
        cogs = cogs_per_sale.get(sale.id, 0.0)
        prof = rev - cogs
        margin = (prof / rev) * 100 if rev > 0 else 0.0
        detailed_sales.append({
            'sale': sale,
            'revenue': rev,
            'cogs': cogs,
            'profit': prof,
            'margin': margin
        })

    return render_template(
        'admin/reports/income.html',
        detailed_sales=detailed_sales,
        pagination=pagination,
        total_revenue=total_revenue,
        total_cogs=total_cogs,
        net_profit=net_profit,
        profit_margin=profit_margin,
        shop_breakdown=shop_breakdown,
        category_breakdown=category_breakdown,
        from_date=from_date_str,
        to_date=to_date_str
    )


# --- EXPORT REPORTS TO EXCEL ---
@dashboard_bp.route('/admin/reports/sales/export')
@login_required
def export_sales_report():
    import io
    from datetime import datetime, timedelta
    from sqlalchemy import func
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file

    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    branch_id = request.args.get('branch_id', type=int)
    category_id = request.args.get('category_id', type=int)
    product_id = request.args.get('product_id', type=int)

    # Date parsing (same as sales_report)
    if not from_date_str:
        from_date_dt = datetime.now() - timedelta(days=30)
    else:
        try:
            from_date_dt = datetime.fromisoformat(from_date_str)
        except ValueError:
            from_date_dt = datetime.now() - timedelta(days=30)

    if not to_date_str:
        to_date_dt = datetime.now()
    else:
        try:
            to_date_dt = datetime.fromisoformat(to_date_str)
        except ValueError:
            to_date_dt = datetime.now()

    # Query matching sales (scoped to logged-in user)
    sales_query = Sale.query.filter(
        Sale.sale_date >= from_date_dt,
        Sale.sale_date <= to_date_dt,
        Sale.user_id == current_user.id
    )

    # Apply additional filters
    if branch_id:
        sales_query = sales_query.filter(Sale.shop_id == branch_id)
    if category_id:
        sale_ids_query = db.session.query(SaleItem.sale_id).join(Product).filter(
            Product.category_id == category_id,
            SaleItem.user_id == current_user.id
        )
        sales_query = sales_query.filter(Sale.id.in_(sale_ids_query))
    if product_id:
        sale_ids_query = db.session.query(SaleItem.sale_id).filter(
            SaleItem.product_id == product_id,
            SaleItem.user_id == current_user.id
        )
        sales_query = sales_query.filter(Sale.id.in_(sale_ids_query))

    sales_subquery = sales_query.with_entities(Sale.id)

    # Calculate Aggregates
    total_sales_val = db.session.query(func.sum(Sale.total)).filter(Sale.id.in_(sales_subquery)).scalar() or 0.0
    total_sales_val = float(total_sales_val)
    total_discount_val = db.session.query(func.sum(Sale.discount_amount)).filter(Sale.id.in_(sales_subquery)).scalar() or 0.0
    total_discount_val = float(total_discount_val)
    total_transactions = sales_query.count()

    # Sales by Shop
    shop_sales = db.session.query(
        Shop.name,
        func.sum(Sale.total).label('shop_total')
    ).join(Sale, Sale.shop_id == Shop.id)\
     .filter(Sale.id.in_(sales_subquery), Shop.user_id == current_user.id)\
     .group_by(Shop.id).all()

    # Sales by Payment Method
    pm_sales = db.session.query(
        PaymentMethod.name,
        func.sum(Sale.total).label('pm_total')
    ).join(Sale, Sale.payment_method_id == PaymentMethod.id)\
     .filter(Sale.id.in_(sales_subquery), PaymentMethod.user_id == current_user.id)\
     .group_by(PaymentMethod.id).all()

    # All sales in detailed order
    all_sales = sales_query.order_by(Sale.sale_date.desc()).all()

    # Create Workbook
    wb = Workbook()
    
    # Fonts, alignments, fills
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    section_font = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    # --- Sheet 1: Summary ---
    ws1 = wb.active
    ws1.title = "Sales Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    # Report Title
    ws1['A1'] = "Sales Performance Report"
    ws1['A1'].font = title_font
    
    ws1['A2'] = f"Date Range: {from_date_dt.strftime('%Y-%m-%d %H:%M')} to {to_date_dt.strftime('%Y-%m-%d %H:%M')}"
    ws1['A2'].font = regular_font
    ws1['A3'] = f"Generated By: {current_user.username} | Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1['A3'].font = regular_font
    
    # Summary Metrics Block
    ws1['A5'] = "Summary Metrics"
    ws1['A5'].font = section_font
    
    metrics_headers = ["Metric", "Value"]
    for col_idx, h in enumerate(metrics_headers, start=1):
        cell = ws1.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")
        
    metrics_data = [
        ("Total Net Sales", total_sales_val),
        ("Total Transactions", total_transactions),
        ("Total Discounts Applied", total_discount_val)
    ]
    for row_idx, (m, val) in enumerate(metrics_data, start=7):
        ws1.cell(row=row_idx, column=1, value=m).font = regular_font
        cell_val = ws1.cell(row=row_idx, column=2, value=val)
        cell_val.font = bold_font if row_idx == 7 else regular_font
        cell_val.alignment = Alignment(horizontal="right")
        if row_idx in (7, 9):
            cell_val.number_format = "$#,##0.00"
        else:
            cell_val.number_format = "#,##0"
            
    # Add borders to metrics table
    for r in range(6, 10):
        for c in range(1, 3):
            ws1.cell(row=r, column=c).border = thin_border
            
    # Shop Location Table
    start_r = 12
    ws1.cell(row=start_r, column=1, value="Sales by Shop Branch").font = section_font
    
    headers_shop = ["Shop Location", "Revenue"]
    for col_idx, h in enumerate(headers_shop, start=1):
        cell = ws1.cell(row=start_r+1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")
        
    curr_r = start_r + 2
    for name, total in shop_sales:
        ws1.cell(row=curr_r, column=1, value=name).font = regular_font
        cell_val = ws1.cell(row=curr_r, column=2, value=float(total))
        cell_val.font = regular_font
        cell_val.alignment = Alignment(horizontal="right")
        cell_val.number_format = "$#,##0.00"
        curr_r += 1
        
    for r in range(start_r+1, curr_r):
        for c in range(1, 3):
            ws1.cell(row=r, column=c).border = thin_border

    # Payment Method Table
    start_r = curr_r + 2
    ws1.cell(row=start_r, column=1, value="Sales by Payment Method").font = section_font
    
    headers_pm = ["Payment Type", "Revenue"]
    for col_idx, h in enumerate(headers_pm, start=1):
        cell = ws1.cell(row=start_r+1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")
        
    curr_r = start_r + 2
    for name, total in pm_sales:
        ws1.cell(row=curr_r, column=1, value=name).font = regular_font
        cell_val = ws1.cell(row=curr_r, column=2, value=float(total))
        cell_val.font = regular_font
        cell_val.alignment = Alignment(horizontal="right")
        cell_val.number_format = "$#,##0.00"
        curr_r += 1
        
    for r in range(start_r+1, curr_r):
        for c in range(1, 3):
            ws1.cell(row=r, column=c).border = thin_border

    # Auto-adjust column widths for Sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # --- Sheet 2: Detailed Transactions ---
    ws2 = wb.create_sheet(title="Transaction Details")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2['A1'] = "Detailed Sales Transaction Log"
    ws2['A1'].font = title_font
    ws2['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws2['A2'].font = regular_font

    headers_details = [
        "Sale ID", "Date-Time", "Shop Location", "Salesperson", 
        "Payment Type", "Discount %", "Discount Amount", "Net Total", "Amount Paid"
    ]
    
    for col_idx, h in enumerate(headers_details, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        if col_idx in (1, 6, 7, 8, 9):
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left")
            
    curr_r = 5
    for sale in all_sales:
        ws2.cell(row=curr_r, column=1, value=sale.id).alignment = Alignment(horizontal="right")
        ws2.cell(row=curr_r, column=2, value=sale.sale_date.strftime('%Y-%m-%d %H:%M:%S'))
        ws2.cell(row=curr_r, column=3, value=sale.shop.name)
        ws2.cell(row=curr_r, column=4, value=sale.user.username)
        ws2.cell(row=curr_r, column=5, value=sale.payment_method.name)
        
        disc_pct_cell = ws2.cell(row=curr_r, column=6, value=sale.discount_pct)
        disc_pct_cell.alignment = Alignment(horizontal="right")
        disc_pct_cell.number_format = "0"
        
        disc_amt_cell = ws2.cell(row=curr_r, column=7, value=float(sale.discount_amount))
        disc_amt_cell.alignment = Alignment(horizontal="right")
        disc_amt_cell.number_format = "$#,##0.00"
        
        total_cell = ws2.cell(row=curr_r, column=8, value=float(sale.total))
        total_cell.font = bold_font
        total_cell.alignment = Alignment(horizontal="right")
        total_cell.number_format = "$#,##0.00"
        
        paid_cell = ws2.cell(row=curr_r, column=9, value=float(sale.paid_amount))
        paid_cell.alignment = Alignment(horizontal="right")
        paid_cell.number_format = "$#,##0.00"
        
        for c in range(1, 10):
            cell = ws2.cell(row=curr_r, column=c)
            cell.border = thin_border
            if c not in (1, 6, 7, 8, 9):
                cell.font = regular_font
                
        curr_r += 1

    # Auto-adjust column widths for Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to Buffer
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        file_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@dashboard_bp.route('/admin/reports/income/export')
@login_required
def export_income_report():
    import io
    from datetime import datetime, timedelta
    from sqlalchemy import func
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file

    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')

    # Date parsing
    if not from_date_str:
        from_date_dt = datetime.now() - timedelta(days=30)
    else:
        try:
            from_date_dt = datetime.fromisoformat(from_date_str)
        except ValueError:
            from_date_dt = datetime.now() - timedelta(days=30)

    if not to_date_str:
        to_date_dt = datetime.now()
    else:
        try:
            to_date_dt = datetime.fromisoformat(to_date_str)
        except ValueError:
            to_date_dt = datetime.now()

    # Query matching sales (scoped to current user)
    sales_query = Sale.query.filter(
        Sale.sale_date >= from_date_dt,
        Sale.sale_date <= to_date_dt,
        Sale.user_id == current_user.id
    )
    sales_subquery = sales_query.with_entities(Sale.id)

    # Calculate Aggregates
    total_revenue = db.session.query(func.sum(Sale.total)).filter(Sale.id.in_(sales_subquery)).scalar() or 0.0
    total_revenue = float(total_revenue)
    
    total_cogs = db.session.query(
        func.sum(SaleItem.qty * func.coalesce(Product.cost, 0.0))
    ).join(Product, SaleItem.product_id == Product.id)\
     .filter(SaleItem.sale_id.in_(sales_subquery), Product.user_id == current_user.id, SaleItem.user_id == current_user.id).scalar() or 0.0
    total_cogs = float(total_cogs)

    net_profit = total_revenue - total_cogs
    profit_margin = (net_profit / total_revenue) * 100 if total_revenue > 0 else 0.0

    # Sales by Shop
    shop_revenues = db.session.query(
        Shop.id,
        Shop.name,
        func.sum(Sale.total).label('revenue')
    ).join(Sale, Sale.shop_id == Shop.id)\
     .filter(Sale.id.in_(sales_subquery), Shop.user_id == current_user.id)\
     .group_by(Shop.id).all()

    shop_cogs = db.session.query(
        Sale.shop_id,
        func.sum(SaleItem.qty * func.coalesce(Product.cost, 0.0)).label('cogs')
    ).join(SaleItem, SaleItem.sale_id == Sale.id)\
     .join(Product, SaleItem.product_id == Product.id)\
     .filter(Sale.id.in_(sales_subquery), Product.user_id == current_user.id, SaleItem.user_id == current_user.id)\
     .group_by(Sale.shop_id).all()

    cogs_map = {row.shop_id: float(row.cogs or 0.0) for row in shop_cogs}
    shop_breakdown = []
    for row in shop_revenues:
        rev = float(row.revenue or 0.0)
        cogs = cogs_map.get(row.id, 0.0)
        prof = rev - cogs
        margin = (prof / rev) * 100 if rev > 0 else 0.0
        shop_breakdown.append({
            'name': row.name,
            'revenue': rev,
            'cogs': cogs,
            'profit': prof,
            'margin': margin
        })

    # Sales by Category
    category_data = db.session.query(
        Category.name,
        func.sum(SaleItem.qty * SaleItem.price * (1 - func.coalesce(Sale.discount_pct, 0) / 100.0)).label('revenue'),
        func.sum(SaleItem.qty * func.coalesce(Product.cost, 0.0)).label('cogs')
    ).select_from(SaleItem)\
     .join(Sale, SaleItem.sale_id == Sale.id)\
     .join(Product, SaleItem.product_id == Product.id)\
     .join(Category, Product.category_id == Category.id)\
     .filter(Sale.id.in_(sales_subquery), Product.user_id == current_user.id, Category.user_id == current_user.id, SaleItem.user_id == current_user.id)\
     .group_by(Category.id).all()

    category_breakdown = []
    for row in category_data:
        rev = float(row.revenue or 0.0)
        cogs = float(row.cogs or 0.0)
        prof = rev - cogs
        margin = (prof / rev) * 100 if rev > 0 else 0.0
        category_breakdown.append({
            'name': row.name,
            'revenue': rev,
            'cogs': cogs,
            'profit': prof,
            'margin': margin
        })

    # All matching sales detailed
    all_sales = sales_query.order_by(Sale.sale_date.desc()).all()
    sale_ids = [s.id for s in all_sales]
    cogs_per_sale = {}
    if sale_ids:
        cogs_rows = db.session.query(
            SaleItem.sale_id,
            func.sum(SaleItem.qty * func.coalesce(Product.cost, 0.0)).label('cogs')
        ).join(Product, SaleItem.product_id == Product.id)\
         .filter(SaleItem.sale_id.in_(sale_ids), Product.user_id == current_user.id, SaleItem.user_id == current_user.id)\
         .group_by(SaleItem.sale_id).all()
        cogs_per_sale = {row.sale_id: float(row.cogs or 0.0) for row in cogs_rows}

    # Create Workbook
    wb = Workbook()
    
    title_font = Font(name="Calibri", size=16, bold=True, color="2E75B6")
    section_font = Font(name="Calibri", size=12, bold=True, color="2E75B6")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    # --- Sheet 1: Profitability Summary ---
    ws1 = wb.active
    ws1.title = "Profitability Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1['A1'] = "Income & Profitability Report"
    ws1['A1'].font = title_font
    
    ws1['A2'] = f"Date Range: {from_date_dt.strftime('%Y-%m-%d %H:%M')} to {to_date_dt.strftime('%Y-%m-%d %H:%M')}"
    ws1['A2'].font = regular_font
    ws1['A3'] = f"Generated By: {current_user.username} | Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1['A3'].font = regular_font

    ws1['A5'] = "Summary Metrics"
    ws1['A5'].font = section_font

    metrics_headers = ["Metric", "Value"]
    for col_idx, h in enumerate(metrics_headers, start=1):
        cell = ws1.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")
        
    metrics_data = [
        ("Total Revenue (Net Sales)", total_revenue),
        ("Cost of Goods Sold (COGS)", total_cogs),
        ("Net Income (Profit)", net_profit),
        ("Profit Margin", profit_margin)
    ]
    for row_idx, (m, val) in enumerate(metrics_data, start=7):
        ws1.cell(row=row_idx, column=1, value=m).font = regular_font
        cell_val = ws1.cell(row=row_idx, column=2, value=val)
        cell_val.font = bold_font if row_idx == 9 else regular_font
        cell_val.alignment = Alignment(horizontal="right")
        if row_idx == 10:
            cell_val.number_format = "0.00\"%\""
        else:
            cell_val.number_format = "$#,##0.00"
            
    for r in range(6, 11):
        for c in range(1, 3):
            ws1.cell(row=r, column=c).border = thin_border

    # Profit by Shop Location
    start_r = 13
    ws1.cell(row=start_r, column=1, value="Profit by Shop Location").font = section_font
    
    headers_shop = ["Shop Location", "Revenue", "COGS", "Net Profit", "Margin"]
    for col_idx, h in enumerate(headers_shop, start=1):
        cell = ws1.cell(row=start_r+1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")
        
    curr_r = start_r + 2
    for item in shop_breakdown:
        ws1.cell(row=curr_r, column=1, value=item['name']).font = regular_font
        
        c2 = ws1.cell(row=curr_r, column=2, value=item['revenue'])
        c2.number_format = "$#,##0.00"
        c2.alignment = Alignment(horizontal="right")
        
        c3 = ws1.cell(row=curr_r, column=3, value=item['cogs'])
        c3.number_format = "$#,##0.00"
        c3.alignment = Alignment(horizontal="right")
        
        c4 = ws1.cell(row=curr_r, column=4, value=item['profit'])
        c4.font = bold_font
        c4.number_format = "$#,##0.00"
        c4.alignment = Alignment(horizontal="right")
        
        c5 = ws1.cell(row=curr_r, column=5, value=item['margin'])
        c5.number_format = "0.00\"%\""
        c5.alignment = Alignment(horizontal="right")
        
        for c in range(1, 6):
            ws1.cell(row=curr_r, column=c).font = regular_font if c != 4 else bold_font
            
        curr_r += 1
        
    for r in range(start_r+1, curr_r):
        for c in range(1, 6):
            ws1.cell(row=r, column=c).border = thin_border

    # Profit by Category
    start_r = curr_r + 2
    ws1.cell(row=start_r, column=1, value="Profit by Category").font = section_font
    
    headers_cat = ["Category", "Revenue", "COGS", "Net Profit", "Margin"]
    for col_idx, h in enumerate(headers_cat, start=1):
        cell = ws1.cell(row=start_r+1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")
        
    curr_r = start_r + 2
    for item in category_breakdown:
        ws1.cell(row=curr_r, column=1, value=item['name']).font = regular_font
        
        c2 = ws1.cell(row=curr_r, column=2, value=item['revenue'])
        c2.number_format = "$#,##0.00"
        c2.alignment = Alignment(horizontal="right")
        
        c3 = ws1.cell(row=curr_r, column=3, value=item['cogs'])
        c3.number_format = "$#,##0.00"
        c3.alignment = Alignment(horizontal="right")
        
        c4 = ws1.cell(row=curr_r, column=4, value=item['profit'])
        c4.font = bold_font
        c4.number_format = "$#,##0.00"
        c4.alignment = Alignment(horizontal="right")
        
        c5 = ws1.cell(row=curr_r, column=5, value=item['margin'])
        c5.number_format = "0.00\"%\""
        c5.alignment = Alignment(horizontal="right")
        
        for c in range(1, 6):
            ws1.cell(row=curr_r, column=c).font = regular_font if c != 4 else bold_font
            
        curr_r += 1
        
    for r in range(start_r+1, curr_r):
        for c in range(1, 6):
            ws1.cell(row=r, column=c).border = thin_border

    # Auto-adjust column widths for Sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # --- Sheet 2: Detailed Profit Margins ---
    ws2 = wb.create_sheet(title="Transaction Profitability")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2['A1'] = "Transaction Profitability Details"
    ws2['A1'].font = title_font
    ws2['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws2['A2'].font = regular_font

    headers_details = [
        "Sale ID", "Date-Time", "Shop Location", "Revenue", "COGS", "Net Profit", "Profit Margin"
    ]
    
    for col_idx, h in enumerate(headers_details, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        if col_idx in (1, 4, 5, 6, 7):
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left")
            
    curr_r = 5
    for sale in all_sales:
        ws2.cell(row=curr_r, column=1, value=sale.id).alignment = Alignment(horizontal="right")
        ws2.cell(row=curr_r, column=2, value=sale.sale_date.strftime('%Y-%m-%d %H:%M:%S'))
        ws2.cell(row=curr_r, column=3, value=sale.shop.name)
        
        rev = float(sale.total)
        cogs = cogs_per_sale.get(sale.id, 0.0)
        prof = rev - cogs
        margin = (prof / rev) * 100 if rev > 0 else 0.0
        
        c4 = ws2.cell(row=curr_r, column=4, value=rev)
        c4.alignment = Alignment(horizontal="right")
        c4.number_format = "$#,##0.00"
        
        c5 = ws2.cell(row=curr_r, column=5, value=cogs)
        c5.alignment = Alignment(horizontal="right")
        c5.number_format = "$#,##0.00"
        
        c6 = ws2.cell(row=curr_r, column=6, value=prof)
        c6.font = bold_font
        c6.alignment = Alignment(horizontal="right")
        c6.number_format = "$#,##0.00"
        
        c7 = ws2.cell(row=curr_r, column=7, value=margin)
        c7.alignment = Alignment(horizontal="right")
        c7.number_format = "0.00\"%\""
        
        for c in range(1, 8):
            cell = ws2.cell(row=curr_r, column=c)
            cell.border = thin_border
            if c not in (1, 4, 5, 6, 7):
                cell.font = regular_font
                
        curr_r += 1

    # Auto-adjust column widths for Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to Buffer
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"income_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        file_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

