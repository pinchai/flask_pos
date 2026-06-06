import sys
import io
from app import app
from extensions import db
from models import Shop, Category, Product
from openpyxl import load_workbook

def test_excel_exports():
    print("Starting Excel Export verification tests...")
    client = app.test_client()

    # 1. Login
    print("Logging in as admin/admin...")
    client.post('/login', data={'username': 'admin', 'password': 'admin'}, follow_redirects=True)

    # Fetch filter entities within app context
    with app.app_context():
        shop = Shop.query.first()
        category = Category.query.first()
        product = Product.query.first()
        
        shop_id = shop.id if shop else None
        category_id = category.id if category else None
        product_id = product.id if product else None

    print(f"Using filter IDs: branch_id={shop_id}, category_id={category_id}, product_id={product_id}")

    # 2. Test Sales Export (without filters)
    print("Testing /admin/reports/sales/export without filters...")
    response = client.get('/admin/reports/sales/export')
    if response.status_code != 200:
        print(f"FAIL: Sales Excel export failed with status code {response.status_code}.")
        sys.exit(1)
        
    print("SUCCESS: Sales Excel export downloaded.")
    
    # Try parsing downloaded stream with openpyxl
    try:
        wb = load_workbook(filename=io.BytesIO(response.data))
        print("SUCCESS: Sales Excel file parsed successfully.")
        print(f"Sheets: {wb.sheetnames}")
        assert "Sales Summary" in wb.sheetnames, "Sales Summary sheet missing"
        assert "Transaction Details" in wb.sheetnames, "Transaction Details sheet missing"
        
        ws1 = wb["Sales Summary"]
        print(f"ws1 A1 value: '{ws1['A1'].value}'")
        assert ws1['A1'].value == "Sales Performance Report", "Incorrect sheet title"
    except Exception as e:
        print(f"FAIL: Sales Excel file parsing failed: {e}")
        sys.exit(1)

    # 3. Test Sales Export (with filters)
    print("Testing /admin/reports/sales/export with branch, category, and product filters...")
    query_params = f"?branch_id={shop_id}&category_id={category_id}&product_id={product_id}"
    response = client.get(f'/admin/reports/sales/export{query_params}')
    if response.status_code != 200:
        print(f"FAIL: Sales Excel export with filters failed with status code {response.status_code}.")
        sys.exit(1)
        
    print("SUCCESS: Sales Excel export with filters downloaded.")
    
    try:
        wb = load_workbook(filename=io.BytesIO(response.data))
        print("SUCCESS: Sales Excel file with filters parsed successfully.")
        assert "Sales Summary" in wb.sheetnames, "Sales Summary sheet missing"
    except Exception as e:
        print(f"FAIL: Sales Excel file with filters parsing failed: {e}")
        sys.exit(1)

    # 4. Test Income Export
    print("\nTesting /admin/reports/income/export...")
    response = client.get('/admin/reports/income/export')
    if response.status_code != 200:
        print(f"FAIL: Income Excel export failed with status code {response.status_code}.")
        sys.exit(1)
        
    print("SUCCESS: Income Excel export downloaded.")
    
    try:
        wb = load_workbook(filename=io.BytesIO(response.data))
        print("SUCCESS: Income Excel file parsed successfully.")
        print(f"Sheets: {wb.sheetnames}")
        assert "Profitability Summary" in wb.sheetnames, "Profitability Summary sheet missing"
    except Exception as e:
        print(f"FAIL: Income Excel file parsing failed: {e}")
        sys.exit(1)

    # 5. Verify Student Isolation in Excel Exports
    print("\nTesting user isolation in Excel exports for student user...")
    client_student = app.test_client()
    client_student.post('/login', data={'username': 'student', 'password': 'student'}, follow_redirects=True)

    with app.app_context():
        from models import User, Sale
        student_user = User.query.filter_by(username='student').first()
        student_sales_count = Sale.query.filter_by(user_id=student_user.id).count()
        student_sales_total = db.session.query(db.func.sum(Sale.total)).filter(
            Sale.user_id == student_user.id
        ).scalar() or 0.0
        student_sales_total = float(student_sales_total)

    # Validate Sales Export
    print("Verifying student Sales Excel export...")
    resp_sales = client_student.get('/admin/reports/sales/export')
    assert resp_sales.status_code == 200, "Student Sales Excel export failed"
    wb_sales = load_workbook(filename=io.BytesIO(resp_sales.data))
    ws_sales = wb_sales["Sales Summary"]
    
    # Row 7: Total Net Sales, Row 8: Total Transactions
    excel_sales_total = float(ws_sales['B7'].value or 0.0)
    excel_sales_count = int(ws_sales['B8'].value or 0)
    
    print(f"Student Excel Sales Total: {excel_sales_total}, Count: {excel_sales_count}")
    print(f"Student Expected Sales Total: {student_sales_total}, Count: {student_sales_count}")
    
    assert excel_sales_count == student_sales_count, f"Student Sales Export count mismatch: {excel_sales_count} vs {student_sales_count}"
    assert abs(excel_sales_total - student_sales_total) < 1e-2, f"Student Sales Export total mismatch: {excel_sales_total} vs {student_sales_total}"
    print("SUCCESS: Student Sales Excel export isolated correctly.")

    # Validate Income Export
    print("Verifying student Income Excel export...")
    resp_income = client_student.get('/admin/reports/income/export')
    assert resp_income.status_code == 200, "Student Income Excel export failed"
    wb_income = load_workbook(filename=io.BytesIO(resp_income.data))
    ws_income = wb_income["Profitability Summary"]
    
    # Row 7: Total Revenue (Net Sales)
    excel_income_total = float(ws_income['B7'].value or 0.0)
    print(f"Student Excel Income Total: {excel_income_total}")
    assert abs(excel_income_total - student_sales_total) < 1e-2, f"Student Income Export total mismatch: {excel_income_total} vs {student_sales_total}"
    print("SUCCESS: Student Income Excel export isolated correctly.")

    print("\nALL EXCEL EXPORT VERIFICATION TESTS PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    test_excel_exports()
